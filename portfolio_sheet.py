"""
Shared Portfolio Scorecard sheet builder — used by both build_workbook.py
(the blank template with example rows) and export_to_excel.py (live data
pulled from a running backend). Kept in one place so the two can't drift:
the layout, formulas, and formatting are defined once here.

Each row in `org_rows` is a dict:
    {
        "name": str, "ein": str,
        "total_revenue": float|None, "total_expenses": float|None,
        "total_assets": float|None, "net_assets": float|None,
        "contributions_revenue": float|None, "program_revenue": float|None,
        "legal": int, "governance": int, "strategy": int, "trackRecord": int,
        "outcomes": int, "leadership": int, "reporting": int,
        "note": str (optional),
    }
Financial fields may be None (org has no ingested snapshot yet) - written
as 0 with a note flagging it, since Excel formulas need numbers, not nulls.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
GROUP_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
GROUP_FILL_RAW = PatternFill("solid", fgColor="4472C4")
GROUP_FILL_CALC = PatternFill("solid", fgColor="808080")
GROUP_FILL_MANUAL = PatternFill("solid", fgColor="2E7D5B")
GROUP_FILL_RESULT = PatternFill("solid", fgColor="B8860B")
TITLE_FONT = Font(name=FONT, bold=True, size=14)
SUBTLE = Font(name=FONT, italic=True, color="666666", size=10)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Organization", None, "raw"),
    ("EIN", None, "raw"),
    ("Total Revenue ($)", "$#,##0", "raw"),
    ("Total Expenses ($)", "$#,##0", "raw"),
    ("Total Assets ($)", "$#,##0", "raw"),
    ("Net Assets ($)", "$#,##0", "raw"),
    ("Contributions ($)", "$#,##0", "raw"),
    ("Program Revenue ($)", "$#,##0", "raw"),
    ("Reserve Months", "0.0", "calc"),
    ("Revenue Concentration", "0.0%", "calc"),
    ("Financial Health (0-100)", "0", "calc"),
    ("Legal & Compliance (0-100)", "0", "manual"),
    ("Governance (0-100)", "0", "manual"),
    ("Strategic Clarity (0-100)", "0", "manual"),
    ("Track Record (0-100)", "0", "manual"),
    ("Outcome Measurement (0-100)", "0", "manual"),
    ("Leadership Stability (0-100)", "0", "manual"),
    ("Reporting Capacity (0-100)", "0", "manual"),
    ("Overall Score", "0.0", "result"),
    ("Status", None, "result"),
]

MANUAL_KEYS = ["legal", "governance", "strategy", "trackRecord", "outcomes", "leadership", "reporting"]
GROUP_ROW = 4
HEADER_ROW = 5
DATA_START = HEADER_ROW + 1
NOTE_COL = 21  # column U, one past Status


def _financial_health_formula(col, r):
    reserve_c, conc_c = f"{col['Reserve Months']}{r}", f"{col['Revenue Concentration']}{r}"
    rev_c, exp_c = f"{col['Total Revenue ($)']}{r}", f"{col['Total Expenses ($)']}{r}"
    # Guard matches scoring.py's score_financial_health(): no revenue AND no
    # expenses on the row means no financial data was ever entered/ingested,
    # so this returns the same neutral 50 default the backend returns —
    # without the guard, zeroed inputs would compute a misleadingly "good"
    # score instead of honestly reporting "no data" (a real bug caught while
    # testing export_to_excel.py against an org with no ingested snapshot).
    computed = (f"ROUND(MIN(100,({reserve_c}/6)*100)*0.4"
                f"+IF({conc_c}<=0.5,100,MAX(0,100-(({conc_c}-0.5)/0.5)*100))*0.35"
                f"+IF({rev_c}>={exp_c},100,MAX(0,100-(({exp_c}-{rev_c})/MAX({rev_c},1))*200))*0.25,0)")
    return f"=IF(AND({rev_c}=0,{exp_c}=0),50,{computed})"


def _write_formula_row(ws, col, r):
    """Writes the calc/result formulas for row r. Assumes raw + manual inputs
    are already in place (or will be typed in later - formulas work either
    way)."""
    exp_c, net_c = f"{col['Total Expenses ($)']}{r}", f"{col['Net Assets ($)']}{r}"
    rev_c = f"{col['Total Revenue ($)']}{r}"
    contrib_c, prog_c = f"{col['Contributions ($)']}{r}", f"{col['Program Revenue ($)']}{r}"

    ws[f"{col['Reserve Months']}{r}"] = f"=IF({exp_c}=0,0,{net_c}/({exp_c}/12))"
    ws[f"{col['Reserve Months']}{r}"].font = BLACK
    ws[f"{col['Revenue Concentration']}{r}"] = f"=IF({rev_c}=0,0,MAX({contrib_c},{prog_c})/{rev_c})"
    ws[f"{col['Revenue Concentration']}{r}"].font = BLACK

    fin_cell = f"{col['Financial Health (0-100)']}{r}"
    ws[fin_cell] = _financial_health_formula(col, r)
    ws[fin_cell].font = BLACK

    legal_c, gov_c, strat_c, track_c, out_c, lead_c, rep_c = (
        f"{col['Legal & Compliance (0-100)']}{r}", f"{col['Governance (0-100)']}{r}",
        f"{col['Strategic Clarity (0-100)']}{r}", f"{col['Track Record (0-100)']}{r}",
        f"{col['Outcome Measurement (0-100)']}{r}", f"{col['Leadership Stability (0-100)']}{r}",
        f"{col['Reporting Capacity (0-100)']}{r}",
    )
    overall_cell = f"{col['Overall Score']}{r}"
    ws[overall_cell] = (f"=ROUND(({legal_c}*15+{fin_cell}*20+{gov_c}*10+{strat_c}*10+{track_c}*15"
                        f"+{out_c}*15+{lead_c}*5+{rep_c}*10)/100,1)")
    ws[overall_cell].font = Font(name=FONT, bold=True)

    status_cell = f"{col['Status']}{r}"
    ws[status_cell] = f'=IF({overall_cell}>=75,"Ready",IF({overall_cell}>=55,"Needs Work","Not Ready"))'
    ws[status_cell].font = Font(name=FONT, bold=True)

    for name, fmt, _group in COLUMNS:
        c = ws[f"{col[name]}{r}"]
        if fmt:
            c.number_format = fmt
        c.border = BORDER


def add_portfolio_scorecard_sheet(wb, org_rows, sheet_name="Portfolio Scorecard",
                                   subtitle=None, blank_rows=18):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    ws["A1"] = "Portfolio Scorecard — full 8-dimension readiness rubric, multiple organizations"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle or (
        "Blue cells are raw inputs. Green cells are manual scores (0-100) for the 7 qualitative "
        "dimensions. Gray and gold cells are formulas that recalculate automatically."
    )
    ws["A2"].font = SUBTLE
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 44
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    col = {}
    for idx, (name, fmt, group) in enumerate(COLUMNS, start=1):
        col[name] = get_column_letter(idx)
        c = ws.cell(row=HEADER_ROW, column=idx, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[HEADER_ROW].height = 34

    group_bands = [
        (1, 2, "Organization", GROUP_FILL_RAW),
        (3, 8, "Raw Financials — input", GROUP_FILL_RAW),
        (9, 11, "Financial Calc — auto", GROUP_FILL_CALC),
        (12, 18, "Manual Dimension Scores (0-100) — input", GROUP_FILL_MANUAL),
        (19, 20, "Result", GROUP_FILL_RESULT),
    ]
    for start, end, label, fill in group_bands:
        ws.merge_cells(start_row=GROUP_ROW, start_column=start, end_row=GROUP_ROW, end_column=end)
        c = ws.cell(row=GROUP_ROW, column=start, value=label)
        c.font = GROUP_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center")

    r = DATA_START
    for row in org_rows:
        ws[f"{col['Organization']}{r}"] = row.get("name", "")
        ws[f"{col['Organization']}{r}"].font = BLUE
        ws[f"{col['EIN']}{r}"] = row.get("ein", "")
        ws[f"{col['EIN']}{r}"].font = BLUE
        for key, name in [
            ("total_revenue", "Total Revenue ($)"), ("total_expenses", "Total Expenses ($)"),
            ("total_assets", "Total Assets ($)"), ("net_assets", "Net Assets ($)"),
            ("contributions_revenue", "Contributions ($)"), ("program_revenue", "Program Revenue ($)"),
        ]:
            val = row.get(key)
            ws[f"{col[name]}{r}"] = val if val is not None else 0
            ws[f"{col[name]}{r}"].font = BLUE
        for key, name in zip(MANUAL_KEYS, [
            "Legal & Compliance (0-100)", "Governance (0-100)", "Strategic Clarity (0-100)",
            "Track Record (0-100)", "Outcome Measurement (0-100)", "Leadership Stability (0-100)",
            "Reporting Capacity (0-100)",
        ]):
            val = row.get(key, 50)
            c = ws[f"{col[name]}{r}"]
            c.value = val
            c.font = Font(name=FONT, color="1B5E20")

        _write_formula_row(ws, col, r)

        note = row.get("note")
        if note:
            nc = ws.cell(row=r, column=NOTE_COL, value=note)
            nc.font = SUBTLE
        r += 1

    blank_end = r + blank_rows
    for br in range(r, blank_end):
        ws[f"{col['Organization']}{br}"] = ""
        ws[f"{col['EIN']}{br}"] = ""
        for name, _fmt, _group in COLUMNS[2:8]:
            ws[f"{col[name]}{br}"] = 0
            ws[f"{col[name]}{br}"].font = BLUE
        for name in ["Legal & Compliance (0-100)", "Governance (0-100)", "Strategic Clarity (0-100)",
                     "Track Record (0-100)", "Outcome Measurement (0-100)", "Leadership Stability (0-100)",
                     "Reporting Capacity (0-100)"]:
            ws[f"{col[name]}{br}"] = 50
            ws[f"{col[name]}{br}"].font = Font(name=FONT, color="1B5E20")
        _write_formula_row(ws, col, br)

    last_row = blank_end - 1
    score_range = f"{col['Overall Score']}{DATA_START}:{col['Overall Score']}{last_row}"
    status_range = f"{col['Status']}{DATA_START}:{col['Status']}{last_row}"
    ws.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=GREEN_FILL))
    ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["55", "74.999"], fill=AMBER_FILL))
    ws.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["55"], fill=RED_FILL))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{col["Status"]}{DATA_START}="Ready"'], fill=GREEN_FILL))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{col["Status"]}{DATA_START}="Needs Work"'], fill=AMBER_FILL))
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'{col["Status"]}{DATA_START}="Not Ready"'], fill=RED_FILL))

    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="100",
                         errorTitle="Invalid score", error="Enter a whole number from 0 to 100.")
    ws.add_data_validation(dv)
    for name in ["Legal & Compliance (0-100)", "Governance (0-100)", "Strategic Clarity (0-100)",
                 "Track Record (0-100)", "Outcome Measurement (0-100)", "Leadership Stability (0-100)",
                 "Reporting Capacity (0-100)"]:
        dv.add(f"{col[name]}{DATA_START}:{col[name]}{last_row}")

    weights_row = blank_end + 1
    ws.cell(row=weights_row, column=1,
            value=("Weights: Legal 15% · Financial 20% · Governance 10% · Strategy 10% · "
                   "Track Record 15% · Outcomes 15% · Leadership 5% · Reporting 10% — "
                   "matches RUBRIC in scoring.py")).font = SUBTLE
    ws.merge_cells(f"A{weights_row}:F{weights_row}")
    ws[f"A{weights_row}"].alignment = Alignment(wrap_text=True)

    widths = {"A": 26, "B": 14, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15,
              "I": 12, "J": 12, "K": 13, "L": 11, "M": 11, "N": 11, "O": 11, "P": 11, "Q": 11, "R": 11,
              "S": 11, "T": 13, "U": 55}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w

    return ws
