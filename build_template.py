"""
Generates a blank Portfolio Scorecard workbook with two worked examples
(Bright Path Youth Alliance, Riverside Community Kitchen — the same sample
orgs used throughout this project) plus reference sheets. Uses
portfolio_sheet.py so the layout is identical to what export_to_excel.py
produces from live data.

Usage:
    python3 build_template.py [output.xlsx]

Note: build in a scratch directory and copy the result where you want it,
rather than repeatedly overwriting a file in place - some mounted/synced
folders don't handle in-place .xlsx rewrites cleanly (this bit us once
during development; see README).
"""
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from portfolio_sheet import (
    add_portfolio_scorecard_sheet, FONT, TITLE_FONT, SUBTLE, HEADER_FONT, HEADER_FILL,
    BLUE, BLACK, YELLOW_FILL, BORDER,
)

LABEL_FONT = Font(name=FONT, bold=True)


def build(out_path="grantpass-financial-workbook.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # portfolio_sheet.add_portfolio_scorecard_sheet creates its own sheet

    # ---------- EIN Input Template ----------
    ws1 = wb.create_sheet("EIN Input Template")
    ws1.sheet_view.showGridLines = False
    ws1["A1"] = "Bulk Load Input — for bulk_ingest.py"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = ("Fill in the EIN column below (required). Name is optional — if left blank, "
                 "bulk_ingest.py uses the name ProPublica has on file. Save this sheet as a CSV "
                 "and run: python3 bulk_ingest.py --token <your-token> --input <this-file>.csv")
    ws1["A2"].font = SUBTLE
    ws1.merge_cells("A2:E2")
    ws1.row_dimensions[2].height = 30
    ws1["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    for i, h in enumerate(["ein", "name", "notes"], start=1):
        c = ws1.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    ws1.cell(row=5, column=1, value="133556677").font = BLUE
    ws1.cell(row=5, column=2, value="Bright Path Youth Alliance").font = BLUE
    ws1.cell(row=5, column=3, value="Example row — edit or delete").font = SUBTLE
    for r in range(5, 31):
        for col in range(1, 4):
            ws1.cell(row=r, column=col).border = BORDER
    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 34
    ws1.column_dimensions["C"].width = 34

    # ---------- Portfolio Scorecard (shared builder) ----------
    org_rows = [
        {
            "name": "Bright Path Youth Alliance", "ein": "133556677",
            "total_revenue": 2415000, "total_expenses": 2260000, "total_assets": 1980000,
            "net_assets": 1120000, "contributions_revenue": 1680000, "program_revenue": 410000,
            "legal": 90, "governance": 85, "strategy": 80, "trackRecord": 88,
            "outcomes": 82, "leadership": 78, "reporting": 92,
            "note": "Financials: real fixture (sample_990.xml / ProPublica). Dimension scores: same values used in the backend test run.",
        },
        {
            "name": "Riverside Community Kitchen", "ein": "042614022",
            "total_revenue": 180000, "total_expenses": 175000, "total_assets": 25000,
            "net_assets": 20500, "contributions_revenue": 140400, "program_revenue": 39600,
            "legal": 55, "governance": 40, "strategy": 55, "trackRecord": 50,
            "outcomes": 45, "leadership": 40, "reporting": 60,
            "note": "Financials: ILLUSTRATIVE — estimated from the org's earlier narrative description, not a filed 990.",
        },
    ]
    add_portfolio_scorecard_sheet(
        wb, org_rows,
        subtitle=("Blue cells are raw inputs. Green cells are manual scores (0-100) for the 7 "
                  "qualitative dimensions — same inputs POST /orgs/{id}/dimensions expects. Gray "
                  "and gold cells are formulas. Two example rows use the same organizations used "
                  "throughout this project."),
    )

    # ---------- Financial Health — How It Works ----------
    ws3 = wb.create_sheet("Financial Health — How It Works")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = "Financial Health Scoring — step-by-step walkthrough"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = ("Same math as the Financial Health formula in Portfolio Scorecard, broken into "
                 "individual steps for one organization. Mirrors score_financial_health() in scoring.py.")
    ws3["A2"].font = SUBTLE
    ws3.merge_cells("A2:D2")
    ws3.row_dimensions[2].height = 30
    ws3["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    rows = [
        ("Organization", "Bright Path Youth Alliance", None), ("EIN", "133556677", None),
        ("Fiscal Year", 2024, None), ("Total Revenue ($)", 2415000, "$#,##0"),
        ("Total Expenses ($)", 2260000, "$#,##0"), ("Total Assets ($)", 1980000, "$#,##0"),
        ("Net Assets ($)", 1120000, "$#,##0"), ("Contributions Revenue ($)", 1680000, "$#,##0"),
        ("Program Revenue ($)", 410000, "$#,##0"),
    ]
    r = 4
    ref = {}
    for label, value, fmt in rows:
        ws3.cell(row=r, column=1, value=label).font = LABEL_FONT
        cell = ws3.cell(row=r, column=2, value=value)
        cell.font = BLUE
        if fmt:
            cell.number_format = fmt
        ref[label] = f"B{r}"
        r += 1
    r += 1
    ws3.cell(row=r, column=1, value="Computed").font = Font(name=FONT, bold=True, size=12)
    r += 1
    rev, exp, net_assets, contrib, prog = (ref["Total Revenue ($)"], ref["Total Expenses ($)"],
                                            ref["Net Assets ($)"], ref["Contributions Revenue ($)"],
                                            ref["Program Revenue ($)"])
    mexp_row = r
    ws3.cell(row=mexp_row, column=1, value="Monthly Expenses ($)").font = LABEL_FONT
    c = ws3.cell(row=mexp_row, column=2, value=f"=({exp})/12"); c.font = BLACK; c.number_format = "$#,##0"
    mexp_ref = f"B{mexp_row}"
    reserve_row = mexp_row + 1
    ws3.cell(row=reserve_row, column=1, value="Reserve Months").font = LABEL_FONT
    c = ws3.cell(row=reserve_row, column=2, value=f"=IF({mexp_ref}=0,0,({net_assets})/{mexp_ref})"); c.font = BLACK; c.number_format = "0.0"
    reserve_ref = f"B{reserve_row}"
    conc_row = reserve_row + 1
    ws3.cell(row=conc_row, column=1, value="Revenue Concentration").font = LABEL_FONT
    c = ws3.cell(row=conc_row, column=2, value=f"=IF({rev}=0,0,MAX({contrib},{prog})/{rev})"); c.font = BLACK; c.number_format = "0.0%"
    conc_ref = f"B{conc_row}"
    rs_row = conc_row + 1
    ws3.cell(row=rs_row, column=1, value="Reserve Score (0-100)").font = LABEL_FONT
    c = ws3.cell(row=rs_row, column=2, value=f"=MIN(100,({reserve_ref}/6)*100)"); c.font = BLACK; c.number_format = "0"
    rs_ref = f"B{rs_row}"
    cs_row = rs_row + 1
    ws3.cell(row=cs_row, column=1, value="Concentration Score (0-100)").font = LABEL_FONT
    c = ws3.cell(row=cs_row, column=2, value=f"=IF({conc_ref}<=0.5,100,MAX(0,100-(({conc_ref}-0.5)/0.5)*100))"); c.font = BLACK; c.number_format = "0"
    cs_ref = f"B{cs_row}"
    ss_row = cs_row + 1
    ws3.cell(row=ss_row, column=1, value="Surplus/Deficit Score (0-100)").font = LABEL_FONT
    c = ws3.cell(row=ss_row, column=2, value=f"=IF({rev}>={exp},100,MAX(0,100-(({exp}-{rev})/MAX({rev},1))*200))"); c.font = BLACK; c.number_format = "0"
    ss_ref = f"B{ss_row}"
    final_row = ss_row + 2
    ws3.cell(row=final_row, column=1, value="Financial Health Score").font = Font(name=FONT, bold=True, size=12)
    c = ws3.cell(row=final_row, column=2, value=f"=ROUND({rs_ref}*0.4+{cs_ref}*0.35+{ss_ref}*0.25,0)")
    c.font = Font(name=FONT, bold=True, size=12); c.number_format = "0"; c.fill = YELLOW_FILL
    note_row = final_row + 2
    ws3.cell(row=note_row, column=1, value="Weights: Reserve 40% / Concentration 35% / Surplus 25% — same as scoring.py").font = SUBTLE
    ws3.merge_cells(f"A{note_row}:D{note_row}")
    ws3[f"A{note_row}"].alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 20

    # ---------- Free vs Candid ----------
    ws4 = wb.create_sheet("Free vs Candid")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "What's free (IRS/ProPublica) vs. what needs a Candid account"
    ws4["A1"].font = TITLE_FONT
    ws4.merge_cells("A1:C1")
    for i, h in enumerate(["Data point", "Free today (bulk_ingest.py)", "Requires a Candid API account"], start=1):
        c = ws4.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
    data = [
        ("Total revenue / expenses / assets", "Yes", ""), ("Net assets / reserves", "Yes", ""),
        ("Contributions vs. program revenue split", "Yes", ""), ("Filing history (multi-year)", "Yes", ""),
        ("Legal name, EIN, state", "Yes", ""), ("Mission summary & logo", "", "Yes"),
        ("Demographics data", "", "Yes"), ("Keyword / cause-area search", "", "Yes"),
        ("Verified nonprofit profile (\"Candid Seal\")", "", "Yes"),
    ]
    rr = 4
    for label, free, candid in data:
        ws4.cell(row=rr, column=1, value=label).font = BLACK
        ws4.cell(row=rr, column=2, value=free).font = BLACK
        ws4.cell(row=rr, column=3, value=candid).font = BLACK
        for col in range(1, 4):
            ws4.cell(row=rr, column=col).border = BORDER
        rr += 1
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 24
    ws4.column_dimensions["C"].width = 28

    wb.save(out_path)
    print("saved", out_path)


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "grantpass-financial-workbook.xlsx")
